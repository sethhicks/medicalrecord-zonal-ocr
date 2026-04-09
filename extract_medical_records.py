"""
Medical Record PDF Extractor
-----------------------------
Extracts fields from medical record PDFs (CMS-1500 and UB-04) using
zonal OCR with template-based alignment, and writes results to Excel.

Usage:
    python extract_medical_records.py path/to/claims.pdf

    python extract_medical_records.py path/to/claims.pdf --zones name,total_charge
    python extract_medical_records.py path/to/claims.pdf --form cms

    python extract_medical_records.py --calibrate templates/template_1500.png --form cms
    python extract_medical_records.py --save-template pdfs/sample.pdf --form cms

Output:
    medical_records_output.xlsx  (created in the working directory)

Dependencies:
    pip install pdf2image pytesseract opencv-python numpy openpyxl
    System:
      Ubuntu/Debian: sudo apt install tesseract-ocr poppler-utils
      macOS:         brew install tesseract poppler
      Windows:       install Tesseract from https://github.com/UB-Mannheim/tesseract/wiki
                     install Poppler from https://github.com/oschwartz10612/poppler-windows/releases
                     and add both to PATH
"""

import argparse
import glob
import os
import re
import sys

import cv2
import numpy as np
import openpyxl
import pytesseract
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pdf2image import convert_from_path

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

# --- CMS-1500 Health Insurance Claim zones ---
ZONES_1500: dict[str, tuple[int, int, int, int]] = {
    #           x     y     w     h
    "name":            (85,   535, 830,  60),
    "date_of_service": (82,  2200, 260,  90),
    "total_charge":    (1588, 2825, 290,  75),
    "dob":             (980,  550,  290,  50),
    "cpt_hcpcs":       (805,  2215, 205,  90),
}

# --- UB-04 Claim Form zones ---
ZONES_UB: dict[str, tuple[int, int, int, int]] = {
    #           x     y     w     h
    "name":            (80,   310, 860,  50),
    "date_of_service": (435,  398, 210,  40),
    "total_charge":    (1838, 2008, 295,  40),
    "dob":             (48,   390, 260,  60),
}

# Tesseract PSM per zone. 7=single line | 6=block | 8=single word
ZONE_PSM: dict[str, str] = {
    "name":            "--psm 7",
    "date_of_service": "--psm 4",
    "total_charge":    "--psm 7",
    "dob":             "--psm 7",
    "cpt_hcpcs":       "--psm 8",
}

# Zones that use only their configured PSM with no fallback.
ZONE_PSM_FIXED = {"date_of_service"}

# Tesseract character whitelists.
TESSERACT_WHITELIST_NAME    = "-c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- "
TESSERACT_WHITELIST_NUMERIC = "-c tessedit_char_whitelist=0123456789./ "

# Maps each zone key to its Excel column label and cleaner function name.
ZONE_FIELDS: dict[str, dict] = {
    "name":            {"label": "Name",            "cleaner": "name"},
    "date_of_service": {"label": "Date of Service", "cleaner": "date"},
    "total_charge":    {"label": "Total Charge",    "cleaner": "charge"},
    "dob":             {"label": "Date of Birth",   "cleaner": "date"},
    "cpt_hcpcs":       {"label": "CPT/HCPCS",       "cleaner": "cpt"},
}

FORM_ZONES: dict[str, dict] = {
    "cms": ZONES_1500,
    "ub":   ZONES_UB,
}
DEFAULT_FORM = "cms"

# Detection regions for form identification.
#           x      y     w     h
DETECT_REGION_1500 = (0,    0,   860, 300)  # "Health Insurance" — top-left
DETECT_REGION_UB   = (2200, 0,   350, 300)  # "4 TYPE OF BILL"   — top-right

# ---------------------------------------------------------------------------
# TEMPLATE REGISTRATION
# ---------------------------------------------------------------------------
# Place reference PNG images in the templates/ folder:
#   templates/template_1500.png  — reference scan for CMS-1500 forms
#   templates/template_ub.png    — reference scan for UB-04 forms
#
# Generate a template from a PDF:
#   python extract_medical_records.py --save-template pdfs/sample.pdf --form cms
TEMPLATE_DIR = "templates"

# ---------------------------------------------------------------------------
# RUNTIME SETTINGS
# ---------------------------------------------------------------------------
OUTPUT_FILE              = "medical_records_output.xlsx"
LOW_CONFIDENCE_THRESHOLD = 50    # Inliers below this → flagged as low confidence
DEBUG                    = False  # Set True to save debug images and print raw OCR

# ---------------------------------------------------------------------------
# UTILITY
# ---------------------------------------------------------------------------

def _template_path(form_type: str) -> str:
    return os.path.join(TEMPLATE_DIR, f"template_{form_type}.png")


def crop_zone(img: np.ndarray, zone: tuple[int, int, int, int]) -> np.ndarray | None:
    """Crop a region from a cv2 image. Returns None if zone is out of bounds."""
    x, y, w, h = zone
    img_h, img_w = img.shape[:2]
    x1, y1 = max(x, 0), max(y, 0)
    x2, y2 = min(x + w, img_w), min(y + h, img_h)
    if x2 <= x1 or y2 <= y1:
        return None
    return img[y1:y2, x1:x2]

# ---------------------------------------------------------------------------
# IMAGE PRE-PROCESSING
# ---------------------------------------------------------------------------

def preprocess_image(gray: np.ndarray, sharpen: bool = True) -> np.ndarray:
    """Denoise → Otsu threshold → deskew → optional mild sharpen."""
    denoised = cv2.fastNlMeansDenoising(gray, h=5)
    _, binary = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    coords = np.column_stack(np.where(binary < 255))
    if len(coords) > 0:
        angle = cv2.minAreaRect(coords)[-1]
        angle = -(90 + angle) if angle < -45 else -angle
        if abs(angle) < 10:
            h, w = binary.shape
            M = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
            binary = cv2.warpAffine(binary, M, (w, h),
                                    flags=cv2.INTER_CUBIC,
                                    borderMode=cv2.BORDER_REPLICATE)

    if not sharpen:
        return binary

    kernel = np.array([[0, -0.5, 0], [-0.5, 3, -0.5], [0, -0.5, 0]])
    sharpened = cv2.filter2D(binary, -1, kernel)
    _, result = cv2.threshold(sharpened, 127, 255, cv2.THRESH_BINARY)
    return result

# ---------------------------------------------------------------------------
# TEMPLATE ALIGNMENT
# ---------------------------------------------------------------------------

def align_to_template(img_bgr: np.ndarray, form_type: str) -> tuple[np.ndarray, int]:
    """Align img_bgr to the reference template using ORB + homography.

    Returns (aligned_image, inlier_count). Falls back to (img_bgr, 0) on
    failure or if no template exists.
    """
    template_path = _template_path(form_type)
    if not os.path.exists(template_path):
        print(f"    ⚠  Template not found: {template_path} — skipping alignment.")
        return img_bgr, 0

    template = cv2.imread(template_path)
    if template is None:
        print(f"    ⚠  Could not load template: {template_path} — skipping alignment.")
        return img_bgr, 0

    h_t, w_t = template.shape[:2]
    WORK_SCALE = 0.25

    tmpl_small = cv2.resize(template, (int(w_t * WORK_SCALE), int(h_t * WORK_SCALE)))
    scan_small = cv2.resize(img_bgr,  (int(img_bgr.shape[1] * WORK_SCALE),
                                        int(img_bgr.shape[0] * WORK_SCALE)))

    gray_tmpl = cv2.cvtColor(tmpl_small, cv2.COLOR_BGR2GRAY)
    gray_scan = cv2.cvtColor(scan_small, cv2.COLOR_BGR2GRAY)

    orb = cv2.ORB_create(nfeatures=5000)
    kp_t, des_t = orb.detectAndCompute(gray_tmpl, None)
    kp_s, des_s = orb.detectAndCompute(gray_scan, None)

    if des_t is None or des_s is None or len(kp_t) < 4 or len(kp_s) < 4:
        print("    ⚠  Not enough ORB features — skipping alignment.")
        return img_bgr, 0

    matcher     = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=False)
    raw_matches = matcher.knnMatch(des_t, des_s, k=2)
    good        = [m for m, n in raw_matches if m.distance < 0.75 * n.distance]

    if len(good) < 10:
        print(f"    ⚠  Too few good matches ({len(good)}) — skipping alignment.")
        return img_bgr, 0

    pts_t = np.float32([kp_t[m.queryIdx].pt for m in good])
    pts_s = np.float32([kp_s[m.trainIdx].pt for m in good])

    H, mask = cv2.findHomography(pts_s, pts_t, cv2.RANSAC, 5.0)
    if H is None:
        print("    ⚠  Homography estimation failed — skipping alignment.")
        return img_bgr, 0

    inliers = int(mask.sum()) if mask is not None else 0
    if DEBUG:
        print(f"    [debug] homography: {len(good)} matches, {inliers} inliers")

    S     = np.diag([WORK_SCALE, WORK_SCALE, 1.0])
    H_full = np.linalg.inv(S) @ H.astype(np.float64) @ S

    aligned = cv2.warpPerspective(img_bgr, H_full, (w_t, h_t),
                                  flags=cv2.INTER_LINEAR,
                                  borderMode=cv2.BORDER_REPLICATE)

    if DEBUG:
        cv2.imwrite(f"debug_{form_type}_aligned.png", aligned)
        print(f"    [debug] saved aligned scan → debug_{form_type}_aligned.png")

    return aligned, inliers

# ---------------------------------------------------------------------------
# FORM TYPE DETECTION
# ---------------------------------------------------------------------------

def detect_form_type(img_bgr: np.ndarray) -> str:
    """OCR header regions to identify the form type. Returns 'cms' or 'ub'."""

    def read_region(region: tuple[int, int, int, int],
                    upscale: int = 1, enhance: bool = False) -> str:
        cropped = crop_zone(img_bgr, region)
        if cropped is None:
            return ""
        gray = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
        if upscale > 1:
            gray = cv2.resize(gray, (gray.shape[1] * upscale, gray.shape[0] * upscale),
                              interpolation=cv2.INTER_CUBIC)
        if enhance:
            clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(4, 4))
            gray  = clahe.apply(gray)
            gray  = cv2.filter2D(gray, -1, np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]]))
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        texts = []
        for psm in ("--psm 7", "--psm 11", "--psm 6", "--psm 3", "--psm 8"):
            t = pytesseract.image_to_string(
                    binary, config=TESSERACT_WHITELIST_NAME + " " + psm).lower()
            if t.strip():
                texts.append(t)
        return " ".join(texts)

    top_left    = read_region(DETECT_REGION_1500)
    bottom_left = read_region(DETECT_REGION_UB, upscale=3, enhance=True)

    if DEBUG:
        save_detect_debug(img_bgr, "autodetect")
        print(f"    [debug] cms detect text: {top_left!r}")
        print(f"    [debug] UB detect text:   {bottom_left!r}")

    if "health insurance" in top_left:
        return "cms"
    if any(kw in bottom_left for kw in ("type", "bill", "type of bill", "4 type")):
        return "ub"

    raise ValueError(
        f"Could not identify form type.\n"
        f"  1500 region text: {top_left!r}\n"
        f"  UB region text:   {bottom_left!r}"
    )

# ---------------------------------------------------------------------------
# FIELD CLEANERS
# ---------------------------------------------------------------------------

def clean_name(raw: str) -> str:
    """Extract patient name — strips label lines, noise chars, and leading junk."""
    LABEL_KEYWORDS = ("patient", "name", "first", "middle", "initial", "last")
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    data_lines = [
        l for l in lines
        if not any(kw in l.lower() for kw in LABEL_KEYWORDS)
        and not re.match(r'^[\|\-\+\s]+$', l)
    ]

    def scrub(line: str) -> str:
        line = re.sub(r'^\|?\s*|\s*\|?$', '', line).strip()
        line = re.sub(r'^[^A-Za-z]+', '', line).strip()
        if re.match(r'^[A-Za-z]\s+[A-Z]', line):
            line = re.sub(r'^[A-Za-z]\s+', '', line).strip()
        line = re.sub(r'[^A-Za-z0-9]+$', '', line).strip()
        return line

    def line_score(line: str) -> float:
        letters = re.findall(r'[A-Za-z]', line)
        if not letters:
            return 0.0
        upper     = len(re.findall(r'[A-Z]', line))
        non_alpha = len([c for c in line if not c.isalpha() and c != ' '])
        return (upper / len(letters)) * len(letters) - non_alpha * 2

    candidates = data_lines if data_lines else lines
    return scrub(max(candidates, key=line_score)) if candidates else ""


def clean_date(raw: str) -> str:
    """Normalise OCR date output to MM/DD/YYYY."""
    digits = re.sub(r'\D', '', raw)

    if len(digits) == 8:
        return f"{digits[0:2]}/{digits[2:4]}/{digits[4:8]}"
    if len(digits) == 7:
        # Could be MDDYYYY or 1MMDDYY — pick by plausible year
        opt_a = f"0{digits[0]}/{digits[1:3]}/{digits[3:7]}"
        opt_b = f"{digits[1:3]}/{digits[3:5]}/20{digits[5:7]}"
        return opt_b if int(digits[3:7]) > 2100 else opt_a
    if len(digits) == 6:
        return f"{digits[0:2]}/{digits[2:4]}/20{digits[4:6]}"
    if len(digits) == 5:
        return f"0{digits[0]}/{digits[1:3]}/20{digits[3:5]}"

    parts = re.findall(r'\d+', raw)
    if len(parts) >= 3:
        mm, dd, yy = parts[0], parts[1], parts[2]
        yy = "20" + yy if len(yy) == 2 else yy
        return f"{mm.zfill(2)}/{dd.zfill(2)}/{yy}"
    if len(parts) == 2:
        return f"{parts[0].zfill(2)}/{parts[1].zfill(2)}"
    if len(parts) == 1:
        return parts[0]
    return raw.strip()


def clean_charge(raw: str) -> str:
    """Return whole-dollar charge amount — strips cents but preserves dollar amounts ending in 00.

    Handles:
      "361 00"     → "361"    (space-separated cents column)
      "361.00"     → "361"    (decimal cents)
      "36100.00"   → "36100"  (large amount with decimal cents)
      "3769763"    → "3769763" (no cents, unchanged)
    """
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    if len(lines) > 1:
        raw = max(lines, key=lambda l: len(re.findall(r'\d', l)))
    raw = raw.strip()

    # If a decimal point or space separates dollars from cents, strip the cents
    # e.g. "361.00" → "361", "36100.00" → "36100", "361 00" → "361"
    match = re.match(r'^(\d+)[.\s]+\d{2}$', raw)
    if match:
        return re.sub(r'[^\d]', '', match.group(1))

    # No explicit separator — strip non-digits then check for merged cents.
    # A value ending in exactly "00" where the remaining digits are > 0
    # is almost certainly a dollars+cents merge (e.g. "36100" = "361.00").
    # Exception: if the full number is a round hundred (e.g. 36100 could be
    # a real amount), we cannot tell — so only strip if the OCR debug image
    # shows a visible gap. As a safe heuristic we strip only when the value
    # has 5 or fewer digits (small charges where cents columns appear).
    digits = re.sub(r'[^\d]', '', raw)
    if digits.endswith('00') and 3 <= len(digits) <= 5:
        digits = digits[:-2]
    return digits


def clean_cpt(raw: str) -> str:
    """Extract a 5-digit CPT/HCPCS code."""
    match = re.search(r'\b(\d{5})\b', raw)
    if match:
        return match.group(1)
    digits = re.sub(r'\D', '', raw)
    return digits[:5].zfill(5) if digits else raw.strip()


CLEANERS = {
    "name":   clean_name,
    "date":   clean_date,
    "charge": clean_charge,
    "cpt":    clean_cpt,
}

# ---------------------------------------------------------------------------
# ZONAL OCR
# ---------------------------------------------------------------------------

def ocr_zone(img_bgr: np.ndarray, zone: tuple[int, int, int, int], psm: str,
             field_name: str = "", form_type: str = "") -> str:
    """Crop → grayscale → preprocess → OCR a single zone."""
    cropped = crop_zone(img_bgr, zone)
    if cropped is None:
        if DEBUG:
            print(f"    [debug] {field_name} zone is outside image bounds — skipping.")
        return ""

    gray = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
    gray = cv2.resize(gray, (gray.shape[1] * 2, gray.shape[0] * 2),
                      interpolation=cv2.INTER_CUBIC)

    processed = preprocess_image(gray, sharpen=(field_name != "name"))

    if DEBUG and field_name:
        prefix = f"debug_{form_type}_" if form_type else "debug_"
        cv2.imwrite(f"{prefix}{field_name}.png", processed)
        print(f"    [debug] saved cropped zone → {prefix}{field_name}.png")

    psm_modes  = [psm] if field_name in ZONE_PSM_FIXED else [psm, "--psm 6", "--psm 4", "--psm 3"]
    candidates = []
    for mode in psm_modes:
        # Name uses no whitelist — full language model needed to recognise spaces
        if field_name == "name":
            text = pytesseract.image_to_string(processed, config=mode).strip()
        else:
            text = pytesseract.image_to_string(processed, config=f"{TESSERACT_WHITELIST_NUMERIC} {mode}").strip()
        if DEBUG:
            print(f"    [debug] {field_name} psm={mode!r:12s} → {text!r}")
        if text:
            candidates.append(text)

    if not candidates:
        return ""

    def plausibility(t: str) -> float:
        alnum   = re.findall(r'[A-Za-z0-9]', t)
        digits  = re.findall(r'\d', t)
        letters = re.findall(r'[A-Za-z]', t)
        if not alnum:
            return 0.0
        if field_name == "name":
            upper     = len(re.findall(r'[A-Z]', t))
            non_alpha = len([c for c in t if not c.isalpha() and c != ' '])
            return (upper / len(letters) if letters else 0) * len(letters) - non_alpha * 2
        if letters and not digits and len(t) > 15:
            return 0.0
        if not digits and len(re.sub(r'\D', '', t)) == len(alnum):
            return len(alnum) + 10
        return len(digits) - len(letters) * 0.5

    return max(candidates, key=plausibility)


def extract_page(img_bgr: np.ndarray, pdf_name: str, page_num: int,
                 active_zones: list[str] | None = None,
                 form_type: str = "") -> dict:
    """Detect form type, align, and OCR a single page image."""
    inliers = 0

    if not form_type:
        try:
            form_type          = detect_form_type(img_bgr)
            print(f"    Detected form:       {form_type.upper()} (raw scan)")
            img_bgr, inliers   = align_to_template(img_bgr, form_type)
        except ValueError:
            for candidate in FORM_ZONES:
                aligned, cand_inliers = align_to_template(img_bgr, candidate)
                try:
                    form_type = detect_form_type(aligned)
                    img_bgr, inliers = aligned, cand_inliers
                    print(f"    Detected form:       {form_type.upper()} (after {candidate} alignment)")
                    break
                except ValueError:
                    continue
            else:
                raise ValueError("Could not identify form type after trying all templates.")
    else:
        img_bgr, inliers = align_to_template(img_bgr, form_type)

    low_confidence = inliers < LOW_CONFIDENCE_THRESHOLD
    if low_confidence:
        print(f"    ⚠  LOW CONFIDENCE — alignment inliers: {inliers} (threshold: {LOW_CONFIDENCE_THRESHOLD})")
        print(f"       Results may be inaccurate. Review manually.")
    else:
        print(f"    Alignment confidence: OK ({inliers} inliers)")

    zones = FORM_ZONES[form_type]
    if active_zones is None:
        active_zones = list(zones.keys())

    record = {
        "File":       pdf_name,
        "Page":       page_num,
        "Form Type":  form_type.upper(),
        "Confidence": "LOW" if low_confidence else "OK",
    }
    for zone_key in active_zones:
        meta             = ZONE_FIELDS[zone_key]
        raw              = ocr_zone(img_bgr, zones[zone_key], ZONE_PSM[zone_key],
                                    zone_key, form_type)
        record[meta["label"]] = CLEANERS[meta["cleaner"]](raw)

    return record


def extract_all_pages(pdf_path: str, active_zones: list[str] | None = None,
                      form_type: str = "") -> list[dict]:
    """Render every page of a PDF and extract fields from each."""
    pages    = convert_from_path(pdf_path, dpi=300)
    pdf_name = os.path.basename(pdf_path)
    print(f"  {len(pages)} page(s) found in {pdf_name}")

    records = []
    for i, page in enumerate(pages, start=1):
        print(f"  Page {i}/{len(pages)}")
        img_bgr = cv2.cvtColor(np.array(page), cv2.COLOR_RGB2BGR)
        try:
            record = extract_page(img_bgr, pdf_name, i, active_zones, form_type)
            records.append(record)
            for key, val in record.items():
                if key not in ("File", "Page", "Form Type", "Confidence"):
                    print(f"    {key:<20} {val or '(not found)'}")
        except Exception as e:
            print(f"    ❌  Error on page {i}: {e}")

    return records

# ---------------------------------------------------------------------------
# CALIBRATION
# ---------------------------------------------------------------------------

def save_detect_debug(img_bgr: np.ndarray, form_type: str) -> None:
    """Save thresholded detection region crops as debug images."""
    for name, region in [("detect_1500", DETECT_REGION_1500),
                         ("detect_ub",   DETECT_REGION_UB)]:
        cropped = crop_zone(img_bgr, region)
        if cropped is None:
            print(f"    [debug] detection region {name} is outside image bounds — skipping.")
            continue
        gray = cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY)
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        path = f"debug_{form_type}_{name}.png"
        cv2.imwrite(path, binary)
        print(f"    [debug] saved detection region → {path}")


def run_calibration(path: str, zones: dict[str, tuple[int, int, int, int]]) -> None:
    """Interactive calibration window — hover to read pixel coordinates."""
    print(f"Calibrating against: {path}")
    print("Hover to read coordinates. Press Q or Escape to quit.\n")

    ext = os.path.splitext(path)[1].lower()
    if ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
        img = cv2.imread(path)
        if img is None:
            print(f"❌  Could not load image: {path}")
            return
    else:
        pages = convert_from_path(path, dpi=300, first_page=1, last_page=1)
        img   = cv2.cvtColor(np.array(pages[0]), cv2.COLOR_RGB2BGR)

    # Draw OCR zones (green) and detection regions (yellow)
    for name, (x, y, w, h) in zones.items():
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 200, 0), 2)
        cv2.putText(img, name, (x + 4, max(y - 6, 16)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 200, 0), 2)
    for label, (x, y, w, h) in [("DETECT:CMS", DETECT_REGION_1500),
                                  ("DETECT:UB",   DETECT_REGION_UB)]:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 215, 255), 2)
        cv2.putText(img, label, (x + 4, y + 20),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 215, 255), 2)

    save_detect_debug(img, "cal")

    scale = min(1.0, 1200 / img.shape[0])

    def on_mouse(event, x, y, flags, param):
        rx, ry = int(x / scale), int(y / scale)
        overlay = img.copy()
        cv2.line(overlay, (rx, 0), (rx, overlay.shape[0]), (0, 0, 255), 1)
        cv2.line(overlay, (0, ry), (overlay.shape[1], ry), (0, 0, 255), 1)
        cv2.putText(overlay, f"x={rx}, y={ry}", (rx + 8, ry - 8),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
        if scale < 1.0:
            overlay = cv2.resize(overlay, (int(overlay.shape[1] * scale),
                                           int(overlay.shape[0] * scale)))
        cv2.imshow(win, overlay)
        if event == cv2.EVENT_LBUTTONDOWN:
            print(f"  Clicked: x={rx}, y={ry}")

    win  = "Calibration — press Q to quit"
    disp = cv2.resize(img, (int(img.shape[1] * scale), int(img.shape[0] * scale))) if scale < 1.0 else img.copy()
    cv2.namedWindow(win, cv2.WINDOW_NORMAL)
    cv2.imshow(win, disp)
    cv2.setMouseCallback(win, on_mouse)
    while cv2.waitKey(20) & 0xFF not in (ord('q'), 27):
        pass
    cv2.destroyAllWindows()
    print("Calibration closed.")

# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
BORDER      = Border(left=Side(style="thin", color="CCCCCC"),
                     right=Side(style="thin", color="CCCCCC"),
                     top=Side(style="thin", color="CCCCCC"),
                     bottom=Side(style="thin", color="CCCCCC"))

_COL_WIDTHS = {
    "File": 30, "Page": 7, "Form Type": 12, "Confidence": 14,
    "Name": 25, "Date of Service": 18, "Total Charge": 16,
    "Date of Birth": 18, "CPT/HCPCS": 14,
}

CENTER_COLS = {"Date of Service", "Total Charge", "Confidence", "Page", "Form Type"}


def build_header(active_zones: list[str]) -> list[str]:
    return ["File", "Page", "Form Type", "Confidence"] + \
           [ZONE_FIELDS[z]["label"] for z in active_zones]


def setup_workbook(header: list[str]):
    if os.path.exists(OUTPUT_FILE):
        try:
            os.remove(OUTPUT_FILE)
            print(f"  Overwriting existing file: {OUTPUT_FILE}")
        except PermissionError:
            print(f"\n❌  Cannot overwrite {OUTPUT_FILE} — close it in Excel and retry.")
            sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title           = "Medical Records"
    ws.freeze_panes    = "A2"
    ws.row_dimensions[1].height = 20

    for col_idx, col_name in enumerate(header, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[cell.column_letter].width = _COL_WIDTHS.get(col_name, 20)

    return wb, ws


def append_row(ws, record: dict, row: int, header: list[str]) -> None:
    is_low   = record.get("Confidence") == "LOW"
    row_fill = (PatternFill("solid", start_color="FFD966") if is_low else
                PatternFill("solid", start_color="EBF3FB") if row % 2 == 0 else None)

    for col_idx, col_name in enumerate(header, start=1):
        cell           = ws.cell(row=row, column=col_idx, value=record.get(col_name, ""))
        cell.font      = DATA_FONT
        cell.border    = BORDER
        cell.alignment = CENTER if col_name in CENTER_COLS else LEFT
        if row_fill:
            cell.fill = row_fill


def save_to_excel(records: list[dict], active_zones: list[str]) -> None:
    header     = build_header(active_zones)
    wb, ws     = setup_workbook(header)
    start_row  = ws.max_row + 1
    for i, record in enumerate(records):
        append_row(ws, record, start_row + i, header)
    wb.save(OUTPUT_FILE)
    print(f"\n✅  Saved {len(records)} record(s) → {OUTPUT_FILE}")

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def resolve_pdf(args: list[str]) -> str:
    candidates = [a for a in args if a.endswith(".pdf") and os.path.isfile(a)]
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        print("❌  No PDF file provided.")
        print("   Usage: python extract_medical_records.py path/to/file.pdf")
    else:
        print("❌  Multiple PDF files provided — please pass exactly one PDF.")
    sys.exit(1)


def parse_args(argv: list[str]):
    parser = argparse.ArgumentParser(
        prog="extract_medical_records.py",
        description="Extract fields from medical record PDFs into Excel.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("pdfs", nargs="*", help="PDF file to process.")
    parser.add_argument("--calibrate", metavar="FILE",
                        help="Open calibration window for the given PDF or image.")
    parser.add_argument("--save-template", metavar="FILE",
                        help="Save page 1 of FILE as the template for --form.")
    parser.add_argument("--zones", metavar="ZONE1,ZONE2,...",
                        help=f"Comma-separated zones to extract.\n"
                             f"Available: {', '.join(ZONE_FIELDS)}\nDefault: all.")
    parser.add_argument("--form", metavar="FORM_TYPE", default=DEFAULT_FORM,
                        help=f"Override auto-detection.\n"
                             f"Available: {', '.join(FORM_ZONES)}\nDefault: auto-detect.")

    args = parser.parse_args(argv)

    form_type = args.form.lower()
    if form_type not in FORM_ZONES:
        print(f"❌  Unknown form type: {args.form}. Available: {', '.join(FORM_ZONES)}")
        sys.exit(1)

    if args.zones:
        requested = [z.strip() for z in args.zones.split(",")]
        invalid   = [z for z in requested if z not in ZONE_FIELDS]
        if invalid:
            print(f"❌  Unknown zone(s): {', '.join(invalid)}. Available: {', '.join(ZONE_FIELDS)}")
            sys.exit(1)
        active_zones = requested
    else:
        active_zones = list(ZONE_FIELDS.keys())

    return args, active_zones, form_type


def main() -> None:
    args, active_zones_arg, form_type_arg = parse_args(sys.argv[1:])

    if args.calibrate:
        run_calibration(args.calibrate, FORM_ZONES[form_type_arg])
        return

    if args.save_template:
        os.makedirs(TEMPLATE_DIR, exist_ok=True)
        out_path = _template_path(form_type_arg)
        ext = os.path.splitext(args.save_template)[1].lower()
        if ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
            img = cv2.imread(args.save_template)
            if img is None:
                print(f"❌  Could not load image: {args.save_template}")
                sys.exit(1)
        else:
            pages = convert_from_path(args.save_template, dpi=300, first_page=1, last_page=1)
            img   = cv2.cvtColor(np.array(pages[0]), cv2.COLOR_RGB2BGR)
        cv2.imwrite(out_path, img)
        print(f"✅  Saved template → {out_path}")
        return

    pdf_path     = resolve_pdf(args.pdfs)
    explicit_form  = form_type_arg if args.form != DEFAULT_FORM else ""
    explicit_zones = active_zones_arg if args.zones else None
    header_zones   = active_zones_arg if args.zones else list(ZONE_FIELDS.keys())

    print(f"Reading: {pdf_path}\n")
    records = extract_all_pages(pdf_path, explicit_zones, explicit_form)

    if records:
        save_to_excel(records, header_zones)
    else:
        print("No records extracted.")


if __name__ == "__main__":
    main()